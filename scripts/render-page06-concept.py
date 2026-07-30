from __future__ import annotations

import argparse
import hashlib
import json
import math
import warnings
from collections import Counter
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.patches import ConnectionPatch, Patch
from openpyxl import load_workbook


ZONE_NAMES = {
    1: "流泥",
    2: "淤泥",
    3: "淤泥质土",
    4: "黏土",
    5: "粉质黏土",
    6: "粉土",
    7: "粉砂–细砂",
    8: "中砂–粗砂",
    9: "砾砂",
}

ZONE_COLORS = {
    1: "#C94332",
    2: "#C8733F",
    3: "#536789",
    4: "#4D9B91",
    5: "#83C8AA",
    6: "#C2A35F",
    7: "#EE9D36",
    8: "#929292",
    9: "#D8D8D8",
}

CURVE_COLORS = {
    "qt": "#C94F4F",
    "rf": "#246B58",
    "u2": "#356FAE",
    "ic": "#F4DC18",
}


def finite(value: object) -> float:
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    return math.nan


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_rows(workbook_path: Path) -> dict[str, np.ndarray]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    raw = list(workbook.worksheets[0].iter_rows(min_row=2, values_only=True))
    result = list(workbook.worksheets[1].iter_rows(min_row=2, values_only=True))
    if len(raw) != len(result) or len(raw) < 2:
        raise ValueError("原始数据与快捷解译结果行数不一致，不能生成样张。")

    depth = np.asarray([finite(row[1]) for row in raw], dtype=float)
    if not np.all(np.isfinite(depth)) or np.any(np.diff(depth) <= 0):
        raise ValueError("深度必须为严格递增的有限数值。")

    zone = np.asarray([
        int(row[11]) if isinstance(row[11], (int, float)) and 1 <= int(row[11]) <= 9 else 0
        for row in result
    ], dtype=int)

    values = {
        "depth": depth,
        "qt": np.asarray([finite(row[4]) / 1000 if math.isfinite(finite(row[4])) else math.nan for row in result]),
        "rf": np.asarray([finite(row[6]) for row in result]),
        "u2": np.asarray([finite(row[4]) for row in raw]),
        "ic": np.asarray([finite(row[10]) for row in result]),
        "zone": zone,
    }
    return values


def break_at_real_gaps(depth: np.ndarray, values: np.ndarray, max_gap: float = 0.05) -> np.ndarray:
    plotted = values.astype(float, copy=True)
    plotted[1:][np.diff(depth) > max_gap] = np.nan
    return plotted


def rolling_majority(depth: np.ndarray, zone: np.ndarray, radius_m: float = 0.50) -> np.ndarray:
    smoothed = np.zeros_like(zone)
    for index, current_depth in enumerate(depth):
        left = int(np.searchsorted(depth, current_depth - radius_m))
        right = int(np.searchsorted(depth, current_depth + radius_m, side="right"))
        counts = Counter(int(value) for value in zone[left:right] if value > 0)
        if counts:
            smoothed[index] = max(counts, key=lambda value: (counts[value], -abs(value - int(zone[index]))))
    return smoothed


def build_reference_layers(
    depth: np.ndarray,
    original_zone: np.ndarray,
    radius_m: float = 0.50,
    minimum_thickness_m: float = 0.50,
    maximum_gap_m: float = 0.05,
) -> list[dict[str, float | int | str]]:
    smoothed = rolling_majority(depth, original_zone, radius_m)
    segments: list[list[int]] = []
    start = 0
    for index in range(1, len(depth)):
        changed = smoothed[index] != smoothed[start]
        separated = depth[index] - depth[index - 1] > maximum_gap_m
        if changed or separated:
            if smoothed[start] > 0:
                segments.append([start, index - 1, int(smoothed[start])])
            start = index
    if smoothed[start] > 0:
        segments.append([start, len(depth) - 1, int(smoothed[start])])

    changed = True
    while changed and len(segments) > 1:
        changed = False
        for index, segment in enumerate(segments):
            start_index, end_index, zone = segment
            if depth[end_index] - depth[start_index] >= minimum_thickness_m:
                continue
            neighbours: list[int] = []
            if index > 0 and depth[start_index] - depth[segments[index - 1][1]] <= maximum_gap_m + 0.03:
                neighbours.append(index - 1)
            if index + 1 < len(segments) and depth[segments[index + 1][0]] - depth[end_index] <= maximum_gap_m + 0.03:
                neighbours.append(index + 1)
            if not neighbours:
                continue
            neighbour_index = min(
                neighbours,
                key=lambda candidate: (
                    abs(segments[candidate][2] - zone),
                    -(depth[segments[candidate][1]] - depth[segments[candidate][0]]),
                ),
            )
            combined_start = min(start_index, segments[neighbour_index][0])
            combined_end = max(end_index, segments[neighbour_index][1])
            counts = Counter(int(value) for value in original_zone[combined_start : combined_end + 1] if value > 0)
            combined_zone = max(counts, key=counts.get) if counts else zone
            insert_at = min(index, neighbour_index)
            for remove_at in sorted([index, neighbour_index], reverse=True):
                segments.pop(remove_at)
            segments.insert(insert_at, [combined_start, combined_end, int(combined_zone)])
            changed = True
            break

    layers: list[dict[str, float | int | str]] = []
    for layer_index, (start_index, end_index, _) in enumerate(segments, start=1):
        evidence = [int(value) for value in original_zone[start_index : end_index + 1] if value > 0]
        if not evidence:
            continue
        counts = Counter(evidence)
        zone = max(counts, key=counts.get)
        confidence = counts[zone] / len(evidence)
        top = float(depth[start_index])
        bottom = float(depth[end_index])
        layers.append({
            "id": f"L{layer_index:03d}",
            "top": top,
            "bottom": bottom,
            "center": (top + bottom) / 2,
            "zone": zone,
            "name": ZONE_NAMES[zone],
            "relative_share": confidence,
            "rows": len(evidence),
        })
    return layers


def spread_label_positions(layers: list[dict[str, float | int | str]], min_depth: float, max_depth: float) -> list[float]:
    if not layers:
        return []
    targets = [float(layer["center"]) for layer in layers]
    min_separation = max(0.72, (max_depth - min_depth) / 58)
    positions = targets.copy()
    for index in range(1, len(positions)):
        positions[index] = max(positions[index], positions[index - 1] + min_separation)
    if positions[-1] > max_depth:
        positions[-1] = max_depth
        for index in range(len(positions) - 2, -1, -1):
            positions[index] = min(positions[index], positions[index + 1] - min_separation)
    if positions[0] < min_depth:
        shift = min_depth - positions[0]
        positions = [position + shift for position in positions]
    return positions


def style_depth_axis(axis: plt.Axes, min_depth: float, max_depth: float, show_depth: bool) -> None:
    axis.set_ylim(max_depth, min_depth)
    major = np.arange(math.ceil(min_depth / 10) * 10, max_depth + 0.1, 10)
    minor = np.arange(math.ceil(min_depth), max_depth + 0.1, 1)
    axis.set_yticks(major)
    axis.set_yticks(minor, minor=True)
    axis.grid(which="major", axis="y", color="#BCC5CA", linewidth=0.75, alpha=0.72)
    axis.grid(which="minor", axis="y", color="#DDE2E5", linewidth=0.35, alpha=0.58)
    axis.grid(which="major", axis="x", color="#C6CDD1", linewidth=0.65, alpha=0.70)
    for spine in axis.spines.values():
        spine.set_color("#22282B")
        spine.set_linewidth(1.05)
    axis.tick_params(axis="both", labelsize=8.5, colors="#202629", length=3.5, width=0.8)
    if show_depth:
        axis.set_ylabel("泥面以下深度 (m)", fontsize=10.5, labelpad=10, weight="semibold")
    else:
        axis.tick_params(axis="y", labelleft=False)


def draw_curve(axis: plt.Axes, depth: np.ndarray, values: np.ndarray, color: str, xlim: tuple[float, float]) -> None:
    plotted = break_at_real_gaps(depth, values)
    clipped = np.clip(plotted, xlim[0], xlim[1])
    axis.plot(clipped, depth, color=color, linewidth=1.15, solid_joinstyle="round", solid_capstyle="round", zorder=4)
    axis.set_xlim(*xlim)


def render(workbook_path: Path, output_path: Path, check_path: Path) -> dict[str, object]:
    values = load_rows(workbook_path)
    depth = values["depth"]
    min_depth = float(depth[0])
    max_depth = float(depth[-1])
    layers = build_reference_layers(depth, values["zone"])
    label_positions = spread_label_positions(layers, min_depth, max_depth)

    plt.rcParams.update({
        "font.family": "sans-serif",
        "font.sans-serif": ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "DejaVu Sans"],
        "svg.fonttype": "none",
        "pdf.fonttype": 42,
        "axes.unicode_minus": False,
        "figure.facecolor": "white",
        "axes.facecolor": "white",
    })

    figure = plt.figure(figsize=(19.2, 10.8), dpi=100, facecolor="white")
    grid = figure.add_gridspec(
        1,
        6,
        left=0.045,
        right=0.986,
        top=0.865,
        bottom=0.135,
        width_ratios=[1.14, 1.04, 1.14, 1.20, 1.18, 1.35],
        wspace=0.40,
    )
    qt_axis = figure.add_subplot(grid[0, 0])
    rf_axis = figure.add_subplot(grid[0, 1], sharey=qt_axis)
    u2_axis = figure.add_subplot(grid[0, 2], sharey=qt_axis)
    ic_axis = figure.add_subplot(grid[0, 3], sharey=qt_axis)
    zone_axis = figure.add_subplot(grid[0, 4], sharey=qt_axis)
    label_axis = figure.add_subplot(grid[0, 5], sharey=qt_axis)

    axes = [qt_axis, rf_axis, u2_axis, ic_axis, zone_axis]
    for index, axis in enumerate(axes):
        style_depth_axis(axis, min_depth, max_depth, index == 0)

    draw_curve(qt_axis, depth, values["qt"], CURVE_COLORS["qt"], (0, 45))
    qt_axis.set_title("修正锥尖阻力 qt", fontsize=11.5, weight="bold", pad=12)
    qt_axis.set_xlabel("修正锥尖阻力\nqt (MPa)", fontsize=9.0, labelpad=7)
    qt_axis.set_xticks([0, 10, 20, 30, 40])

    draw_curve(rf_axis, depth, values["rf"], CURVE_COLORS["rf"], (0, 50))
    rf_axis.set_title("摩阻比 Rf", fontsize=11.5, weight="bold", pad=12)
    rf_axis.set_xlabel("摩阻比\nRf (%)", fontsize=9.0, labelpad=7)
    rf_axis.set_xticks([0, 10, 20, 30, 40, 50])

    draw_curve(u2_axis, depth, values["u2"], CURVE_COLORS["u2"], (-1000, 3500))
    u2_axis.axvline(0, color="#88949A", linewidth=0.75, zorder=2)
    u2_axis.set_title("孔隙水压力 u2", fontsize=11.5, weight="bold", pad=12)
    u2_axis.set_xlabel("孔隙水压力\nu2 (kPa)", fontsize=9.0, labelpad=7)
    u2_axis.set_xticks([-1000, 0, 1000, 2000, 3000])

    ic_bands = [
        (1.00, 1.31, 7),
        (1.31, 2.05, 6),
        (2.05, 2.60, 5),
        (2.60, 2.95, 4),
        (2.95, 3.60, 3),
        (3.60, 4.20, 2),
    ]
    for left, right, zone in ic_bands:
        ic_axis.axvspan(left, right, color=ZONE_COLORS[zone], alpha=0.66, linewidth=0, zorder=0)
        ic_axis.text((left + right) / 2, min_depth + 0.65, f"Z{zone}", ha="center", va="top", fontsize=7.3, color="#263238", weight="bold", zorder=3)
    draw_curve(ic_axis, depth, values["ic"], CURVE_COLORS["ic"], (1.0, 4.2))
    ic_axis.set_title("土体行为类型指数 Ic", fontsize=11.5, weight="bold", pad=12)
    ic_axis.set_xlabel("土体行为类型指数\nIc (-)", fontsize=9.0, labelpad=7)
    ic_axis.set_xticks([1, 2, 3, 4])

    zone_axis.set_xlim(0, 9)
    zone_axis.set_xticks(range(0, 10))
    zone_axis.set_title("归一化土体行为类型", fontsize=11.5, weight="bold", pad=12)
    zone_axis.set_xlabel("JTS 土体行为类型分区\nZone（JTS/T 242—2020）", fontsize=8.5, labelpad=7)
    zone_axis.grid(which="major", axis="x", color="#AEB8BD", linewidth=0.65, alpha=0.72)

    label_axis.set_xlim(0, 1)
    label_axis.set_ylim(max_depth, min_depth)
    label_axis.axis("off")

    for layer, label_depth in zip(layers, label_positions):
        zone = int(layer["zone"])
        top = float(layer["top"])
        bottom = float(layer["bottom"])
        center = float(layer["center"])
        height = max(0.08, bottom - top)
        zone_axis.barh(
            center,
            zone,
            height=height,
            left=0,
            color=ZONE_COLORS[zone],
            edgecolor="white",
            linewidth=0.45,
            align="center",
            zorder=3,
        )
        connection = ConnectionPatch(
            xyA=(zone, center),
            coordsA=zone_axis.transData,
            xyB=(0.025, label_depth),
            coordsB=label_axis.transData,
            color="#71797D",
            linewidth=0.55,
            zorder=2,
            clip_on=False,
        )
        figure.add_artist(connection)
        label_axis.text(
            0.045,
            label_depth,
            f"{layer['id']}  {layer['name']}  {float(layer['relative_share']) * 100:.0f}%",
            fontsize=8.0,
            color="#161A1C",
            va="center",
            ha="left",
            clip_on=False,
        )

    figure.text(0.040, 0.955, "CPT 解译参考地层", fontsize=19, weight="bold", color="#111719", ha="left", va="center")
    figure.text(0.040, 0.925, "SIGS-OGLab support", fontsize=9.5, color="#43545C", ha="left", va="center")
    figure.text(0.962, 0.955, "项目：营口快捷图册  ·  孔位：CPT09-修订", fontsize=9.5, weight="bold", color="#26343A", ha="right", va="center")
    figure.text(0.962, 0.928, f"深度：{min_depth:.2f}–{max_depth:.2f} m  ·  参考层：{len(layers)} 层", fontsize=8.7, color="#67757B", ha="right", va="center")
    figure.add_artist(plt.Line2D([0.04, 0.962], [0.902, 0.902], transform=figure.transFigure, color="#3A454A", linewidth=0.9))

    legend_handles = [Patch(facecolor=ZONE_COLORS[zone], edgecolor="none", label=f"Z{zone} {ZONE_NAMES[zone]}") for zone in range(1, 10)]
    figure.legend(
        handles=legend_handles,
        loc="lower center",
        bbox_to_anchor=(0.50, 0.058),
        ncol=9,
        frameon=False,
        fontsize=8.5,
        handlelength=1.25,
        handleheight=0.9,
        columnspacing=1.25,
        handletextpad=0.45,
    )
    figure.text(
        0.50,
        0.026,
        "参考层由现有 JTS 分类证据按深度窗口确定性汇总；百分比为该层主要类别的相对占比，不代表统计置信度。",
        fontsize=8.0,
        color="#68777D",
        ha="center",
        va="center",
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=100, facecolor="white", edgecolor="none")
    plt.close(figure)

    checks = {
        "schemaVersion": 1,
        "process": "Process118",
        "kind": "page-06-visual-concept",
        "source": {
            "path": workbook_path.as_posix(),
            "sha256": file_sha256(workbook_path),
            "rows": int(len(depth)),
            "depthRangeM": [round(min_depth, 3), round(max_depth, 3)],
        },
        "output": {
            "path": output_path.as_posix(),
            "widthPx": 1920,
            "heightPx": 1080,
        },
        "validCounts": {
            "qt": int(np.count_nonzero(np.isfinite(values["qt"]))),
            "rf": int(np.count_nonzero(np.isfinite(values["rf"]))),
            "u2": int(np.count_nonzero(np.isfinite(values["u2"]))),
            "ic": int(np.count_nonzero(np.isfinite(values["ic"]))),
            "jtsZone": int(np.count_nonzero(values["zone"] > 0)),
        },
        "referenceLayers": {
            "count": len(layers),
            "smoothingRadiusM": 0.50,
            "minimumThicknessM": 0.50,
            "maximumEvidenceGapM": 0.05,
            "percentageMeaning": "dominant-zone relative share within the displayed reference layer",
            "layers": layers,
        },
        "visualChecks": {
            "sharedDepthAxis": True,
            "icThresholdBands": len(ic_bands),
            "externalLayerLabels": len(layers),
            "zoneLegendItems": 9,
            "missingValuesRemainGaps": True,
            "productionCodeChanged": False,
        },
    }
    check_path.write_text(json.dumps(checks, ensure_ascii=False, indent=2), encoding="utf-8")
    return checks


def main() -> None:
    parser = argparse.ArgumentParser(description="Render the standalone Process118 page-six visual concept.")
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--check", required=True, type=Path)
    arguments = parser.parse_args()
    checks = render(arguments.input.resolve(), arguments.output.resolve(), arguments.check.resolve())
    print(json.dumps({
        "output": checks["output"],
        "referenceLayerCount": checks["referenceLayers"]["count"],
        "validCounts": checks["validCounts"],
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
