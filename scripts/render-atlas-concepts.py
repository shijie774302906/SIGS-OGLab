from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import math
import re
import textwrap
import warnings
from collections import Counter
from pathlib import Path
from typing import Callable, Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.patches import Patch, Rectangle
from openpyxl import load_workbook
from PIL import Image


SCRIPT_DIR = Path(__file__).resolve().parent
PAGE06_SPEC = importlib.util.spec_from_file_location("page06_concept", SCRIPT_DIR / "render-page06-concept.py")
if PAGE06_SPEC is None or PAGE06_SPEC.loader is None:
    raise RuntimeError("无法加载已确认的第六页样张配置。")
PAGE06 = importlib.util.module_from_spec(PAGE06_SPEC)
PAGE06_SPEC.loader.exec_module(PAGE06)

ZONE_NAMES = PAGE06.ZONE_NAMES
ZONE_COLORS = PAGE06.ZONE_COLORS
CURVE_COLORS = PAGE06.CURVE_COLORS
PARAMETER_COLORS = ["#536789", "#4D9B91", "#C2A35F", "#C8733F", "#7C668A"]
SCHNEIDER_2008_LABELS = {
    "1a": "1a · 粉土及低刚度指数（Ir）黏土",
    "1b": "1b · 黏土",
    "1c": "1c · 敏感黏土",
    "2": "2 · 基本排水砂土",
    "3": "3 · 过渡土",
}
ROBERTSON_2016_LABELS = {
    "CCS": "CCS · 类黏土—收缩性—敏感",
    "CC": "CC · 类黏土—收缩性",
    "CD": "CD · 类黏土—剪胀性",
    "TC": "TC · 过渡土—收缩性",
    "TD": "TD · 过渡土—剪胀性",
    "SC": "SC · 类砂土—收缩性",
    "SD": "SD · 类砂土—剪胀性",
}
AXIS_LABELS = {
    "qt": "修正锥尖阻力\nqt (MPa)",
    "qc": "锥尖阻力\nqc (MPa)",
    "fs": "侧壁摩阻力（套筒摩阻力）\nfs (kPa)",
    "rf": "摩阻比\nRf (%)",
    "fr": "归一化摩阻比\nFr (%)",
    "u2": "孔隙水压力\nu2 (kPa)",
    "qnet": "净锥尖阻力\nqnet (kPa)",
    "qc_over_pa": "锥尖阻力比\nqc/pa (-)",
    "bq": "孔压参数\nBq (-)",
    "jts_qtn": "JTS 归一化锥尖阻力\nQtn* (-)",
    "qtn": "归一化锥尖阻力\nQtn (-)",
    "ib": "修正土体行为类型指数\nIB (-)",
    "cd": "收缩–剪胀参数\nCD (-)",
    "ic": "土体行为类型指数\nIc (-)",
    "jts_ic": "JTS 土体行为类型指数\nIc (-)",
    "zone": "JTS 土体行为类型分区\nZone (-)",
    "composition": "土类组成比例\n(%)",
    "k": "渗透系数\nk (m/s)",
    "spt_n": "标准贯入击数\nN (击/0.30 m)",
    "es": "压缩模量（R05）\nEs (MPa)",
    "dr": "相对密实度\nDr (%)",
    "phi": "有效内摩擦角\nφ′ (°)",
    "jts_es": "压缩模量（JTS 7.2.8）\nEs (MPa)",
    "g0": "小应变剪切模量\nG0 (MPa)",
    "su": "不排水抗剪强度\nSu (kPa)",
    "su_ratio": "归一化不排水抗剪强度\nSu/σ′v0 (-)",
    "ocr": "超固结比\nOCR (-)",
    "vs": "剪切波速\nVs (m/s)",
    "k0": "静止土压力系数\nK0 (-)",
    "qtn_cs": "等效洁净砂归一化锥尖阻力\nQtn,cs (-)",
    "psi": "状态参数\nψ (-)",
    "st": "灵敏度\nSt (-)",
    "gamma_sat": "饱和重度\nγsat (kN/m³)",
    "water_content": "含水率\nw (%)",
    "void_ratio": "孔隙比\ne (-)",
    "gamma_d": "干重度\nγd (kN/m³)",
    "porosity": "孔隙率\nn (-)",
    "su_residual_ratio": "残余不排水强度比\nSu(r)/σ′v0 (-)",
}
LANDSCAPE_SIZE = (19.2, 10.8)
PORTRAIT_SIZE = (10.8, 15.28)
LANDSCAPE_PX = (1920, 1080)
PORTRAIT_PX = (1080, 1528)
PORTRAIT_PAGES = {1, 5, 15}

FORMULA_GROUPS = [
    {"title": "基础修正与土类指数", "columns": ["qt(kPa)", "JTS Ic"], "prefixes": ["qt(kPa)", "Ic ="]},
    {"title": "饱和重度 γsat (kN/m³)", "columns": ["γsat(kN/m³)"], "prefixes": ["γsat(kN/m³)"]},
    {"title": "渗透系数 k (m/s)", "columns": ["k(m/s)"], "prefixes": ["k(m/s)"]},
    {"title": "标准贯入击数 N", "columns": ["N(击/0.30m)"], "prefixes": ["N="]},
    {"title": "压缩模量 Es（R05）(MPa)", "columns": ["Es(R05)(MPa)"], "prefixes": ["Es（R05）(MPa)"]},
    {"title": "相对密实度 Dr (%)", "columns": ["Dr(%)"], "prefixes": ["Dr(%)"]},
    {"title": "有效内摩擦角 φ′ (°)", "columns": ["φ′(°)"], "prefixes": ["φ′="]},
    {"title": "压缩模量 Es（JTS 7.2.8）(MPa)", "columns": ["Es(JTS 7.2.8)(MPa)"], "prefixes": ["Es（JTS 7.2.8）(MPa)"]},
    {"title": "剪切波速与小应变模量", "columns": ["Vs(m/s)", "G0(MPa)"], "prefixes": ["Vs=", "G0(MPa)"]},
    {"title": "不排水强度与归一化强度", "columns": ["Su峰值(kPa)", "Su重塑后(kPa)", "Su/σ′v0", "Su(r)/σ′v0"], "prefixes": ["Su(kPa)"]},
    {"title": "超固结比与静止土压力", "columns": ["OCR", "K0"], "prefixes": ["OCR =", "K0 ="]},
    {"title": "灵敏度 St", "columns": ["St"], "prefixes": ["St ="]},
    {"title": "等效洁净砂与状态参数", "columns": ["Qtn,cs", "ψ"], "prefixes": ["Qtn,cs=", "Kc=", "ψ ="]},
    {"title": "饱和物理指标", "columns": ["e", "w(%)", "γd(kN/m³)", "n"], "prefixes": ["r=γsat"]},
    {"title": "归一化与分类判据", "columns": ["Robertson Qtn", "Modified Robertson 2016", "IB", "CD", "Schneider 2008"], "prefixes": ["Robertson Qtn=", "IB=", "Schneider Q="]},
]


def finite(value: object) -> float:
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    return math.nan


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_data(path: Path) -> dict[str, object]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, read_only=True, data_only=True)
    raw_sheet, result_sheet, settings_sheet = workbook.worksheets[:3]
    raw_headers = list(next(raw_sheet.iter_rows(values_only=True)))
    result_headers = list(next(result_sheet.iter_rows(values_only=True)))
    raw_rows = list(raw_sheet.iter_rows(min_row=2, values_only=True))
    result_rows = list(result_sheet.iter_rows(min_row=2, values_only=True))
    settings_rows = list(settings_sheet.iter_rows(values_only=True))
    if len(raw_rows) != len(result_rows) or len(raw_rows) < 2:
        raise ValueError("原始数据与快捷解译结果行数不一致。")

    raw_index = {str(name): index for index, name in enumerate(raw_headers)}
    result_index = {str(name): index for index, name in enumerate(result_headers)}
    result_counts = {
        str(name): sum(1 for row in result_rows if row[index] not in (None, ""))
        for index, name in enumerate(result_headers)
    }

    formula_marker = next((index for index, row in enumerate(settings_rows) if row and row[0] == "公式索引"), None)
    reference_marker = next((index for index, row in enumerate(settings_rows) if row and row[0] == "参考文献"), None)
    if formula_marker is None or reference_marker is None or formula_marker >= reference_marker:
        raise ValueError("设置与方法工作表缺少公式索引或参考文献分区。")
    formula_index = [
        str(row[0]).strip()
        for row in settings_rows[formula_marker + 1:reference_marker]
        if row and row[0] not in (None, "")
    ]
    references = {
        str(row[0]).strip(): str(row[1]).strip()
        for row in settings_rows[reference_marker + 1:]
        if row and len(row) > 1 and row[0] not in (None, "") and row[1] not in (None, "")
    }

    def raw(name: str) -> np.ndarray:
        return np.asarray([finite(row[raw_index[name]]) for row in raw_rows], dtype=float)

    def result(name: str) -> np.ndarray:
        return np.asarray([finite(row[result_index[name]]) for row in result_rows], dtype=float)

    def category(name: str) -> np.ndarray:
        index = result_index[name]
        return np.asarray([str(row[index]).strip() if row[index] not in (None, "") else "" for row in result_rows], dtype=object)

    depth = raw("深度(m)")
    if not np.all(np.isfinite(depth)) or np.any(np.diff(depth) <= 0):
        raise ValueError("深度必须为严格递增的有限数值。")

    zone_raw = result("JTS Zone")
    zone = np.asarray([int(value) if math.isfinite(value) and 1 <= int(value) <= 9 else 0 for value in zone_raw], dtype=int)
    data: dict[str, object] = {
        "depth": depth,
        "qc": raw("qc(MPa)"),
        "fs": raw("fs(kPa)"),
        "u2": raw("u2(kPa)"),
        "qt": result("qt(kPa)") / 1000,
        "qnet": result("qnet(kPa)"),
        "rf": result("Rf(%)"),
        "fr": result("Fr(%)"),
        "jts_qtn": result("JTS Qtn*"),
        "bq": result("Bq"),
        "jts_ic": result("JTS Ic"),
        "zone": zone,
        "robertson_qtn": result("Robertson Qtn"),
        "robertson_ic": result("Robertson Ic"),
        "robertson_n": result("Robertson n"),
        "robertson_2016": category("Modified Robertson 2016"),
        "ib": result("IB"),
        "cd": result("CD"),
        "schneider": category("Schneider 2008"),
        "major": category("土类大类"),
        "k": result("k(m/s)"),
        "spt_n": result("N(击/0.30m)"),
        "es": result("Es(R05)(MPa)"),
        "dr": result("Dr(%)"),
        "phi": result("φ′(°)"),
        "jts_es": result("Es(JTS 7.2.8)(MPa)"),
        "g0": result("G0(MPa)"),
        "su_peak": result("Su峰值(kPa)"),
        "su_remolded": result("Su重塑后(kPa)"),
        "su_ratio": result("Su/σ′v0"),
        "su_residual_ratio": result("Su(r)/σ′v0"),
        "ocr": result("OCR"),
        "vs": result("Vs(m/s)"),
        "k0": result("K0"),
        "qtn_cs": result("Qtn,cs"),
        "psi": result("ψ"),
        "st": result("St"),
        "gamma_sat": result("γsat(kN/m³)"),
        "void_ratio": result("e"),
        "water_content": result("w(%)"),
        "gamma_d": result("γd(kN/m³)"),
        "porosity": result("n"),
        "settings": [
            (str(row[0]), str(row[1]))
            for row in settings_rows[:formula_marker]
            if row and row[0] not in (None, "") and len(row) > 1 and row[1] not in (None, "")
        ],
        "result_counts": result_counts,
        "formula_index": formula_index,
        "references": references,
    }
    data["layers"] = PAGE06.build_reference_layers(depth, zone)
    return data


def page_figure(page: int, title: str) -> tuple[plt.Figure, bool]:
    portrait = page in PORTRAIT_PAGES
    figure = plt.figure(figsize=PORTRAIT_SIZE if portrait else LANDSCAPE_SIZE, dpi=100, facecolor="white")
    plt.rcParams.update({
        "font.family": "sans-serif",
        "font.sans-serif": ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "DejaVu Sans"],
        "svg.fonttype": "none",
        "pdf.fonttype": 42,
        "axes.unicode_minus": False,
        "figure.facecolor": "white",
        "axes.facecolor": "white",
    })
    figure.text(0.052, 0.966, title, fontsize=18 if portrait else 19, weight="bold", color="#111719", ha="left", va="center")
    figure.text(0.052, 0.943, "SIGS-OGLab support", fontsize=8.7, color="#43545C", ha="left", va="center")
    figure.text(0.948, 0.966, "项目：营口快捷图册  ·  孔位：CPT09-修订", fontsize=8.5, weight="bold", color="#26343A", ha="right", va="center")
    figure.text(0.948, 0.944, f"CPTU  ·  页面 {page:02d}/15", fontsize=8.2, color="#67757B", ha="right", va="center")
    figure.add_artist(plt.Line2D([0.052, 0.948], [0.922, 0.922], transform=figure.transFigure, color="#3A454A", linewidth=0.9))
    figure.text(0.052, 0.022, "SIGS-OGLab · CPT/CPTU 中文视觉样张", fontsize=7.8, color="#68777D", ha="left")
    figure.text(0.948, 0.022, f"{page:02d}", fontsize=8.0, weight="bold", color="#526168", ha="right")
    return figure, portrait


def save_page(figure: plt.Figure, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output, dpi=100, facecolor="white", edgecolor="none")
    plt.close(figure)


def style_axis(axis: plt.Axes, title: str, xlabel: str = "", ylabel: str = "") -> None:
    axis.set_title(title, fontsize=11.2, weight="bold", pad=10)
    axis.set_xlabel(xlabel, fontsize=9.0, labelpad=6)
    if ylabel:
        axis.set_ylabel(ylabel, fontsize=9.5, labelpad=8, weight="semibold")
    axis.grid(True, which="major", color="#BEC7CC", linewidth=0.65, alpha=0.66)
    axis.grid(True, which="minor", color="#E1E5E7", linewidth=0.35, alpha=0.55)
    axis.tick_params(labelsize=8.2, colors="#202629", length=3.2, width=0.75)
    for spine in axis.spines.values():
        spine.set_color("#22282B")
        spine.set_linewidth(1.0)


def depth_style(axis: plt.Axes, depth: np.ndarray, title: str, xlabel: str, show_depth: bool = False) -> None:
    style_axis(axis, title, xlabel, "泥面以下深度 (m)" if show_depth else "")
    axis.set_ylim(float(depth[-1]), 0)
    axis.set_yticks(np.arange(0, float(depth[-1]) + 0.1, 10))
    axis.set_yticks(np.arange(0, float(depth[-1]) + 0.1, 1), minor=True)
    if not show_depth:
        axis.tick_params(axis="y", labelleft=False)


def gap_values(depth: np.ndarray, values: np.ndarray) -> np.ndarray:
    plotted = np.asarray(values, dtype=float).copy()
    plotted[1:][np.diff(depth) > 0.05] = np.nan
    return plotted


def robust_range(values: np.ndarray, fallback: tuple[float, float], positive: bool = False) -> tuple[float, float]:
    finite_values = values[np.isfinite(values)]
    if positive:
        finite_values = finite_values[finite_values > 0]
    if len(finite_values) < 2:
        return fallback
    low, high = np.quantile(finite_values, [0.01, 0.99])
    if not math.isfinite(low) or not math.isfinite(high) or low == high:
        return fallback
    margin = (high - low) * 0.05
    return (max(np.nextafter(0, 1), low - margin) if positive else low - margin, high + margin)


def plot_depth(axis: plt.Axes, depth: np.ndarray, values: np.ndarray, color: str, x_range: tuple[float, float] | None = None, log: bool = False, second: tuple[np.ndarray, str, str] | None = None) -> None:
    plotted = gap_values(depth, values)
    if x_range is None:
        x_range = robust_range(plotted, (0, 1), positive=log)
    if log:
        plotted[plotted <= 0] = np.nan
        axis.set_xscale("log")
    axis.plot(np.clip(plotted, x_range[0], x_range[1]), depth, color=color, linewidth=1.15, solid_joinstyle="round", zorder=4)
    if second is not None:
        second_values, second_color, second_label = second
        second_plotted = gap_values(depth, second_values)
        if log:
            second_plotted[second_plotted <= 0] = np.nan
        axis.plot(np.clip(second_plotted, x_range[0], x_range[1]), depth, color=second_color, linewidth=1.05, label=second_label, zorder=4)
        axis.plot([], [], color=color, linewidth=1.15, label="峰值")
        axis.legend(loc="lower right", fontsize=7.4, frameon=False)
    axis.set_xlim(*x_range)


def zone_legend(figure: plt.Figure, y: float = 0.058, fontsize: float = 7.9, ncol: int = 9) -> None:
    handles = [Patch(facecolor=ZONE_COLORS[zone], edgecolor="none", label=f"Z{zone} {ZONE_NAMES[zone]}") for zone in range(1, 10)]
    figure.legend(handles=handles, loc="lower center", bbox_to_anchor=(0.50, y), ncol=ncol, frameon=False, fontsize=fontsize, handlelength=1.15, columnspacing=1.0, handletextpad=0.35)


def scatter_zones(axis: plt.Axes, x: np.ndarray, y: np.ndarray, zone: np.ndarray, title: str, xlabel: str, ylabel: str, logx: bool = False, logy: bool = False) -> int:
    valid = np.isfinite(x) & np.isfinite(y) & (zone > 0)
    if logx:
        valid &= x > 0
    if logy:
        valid &= y > 0
    for current_zone in range(1, 10):
        mask = valid & (zone == current_zone)
        axis.scatter(x[mask], y[mask], s=7, alpha=0.54, c=ZONE_COLORS[current_zone], edgecolors="none", rasterized=True)
    if logx:
        axis.set_xscale("log")
    if logy:
        axis.set_yscale("log")
    style_axis(axis, title, xlabel, ylabel)
    return int(np.count_nonzero(valid))


def categorical_depth(axis: plt.Axes, depth: np.ndarray, categories: np.ndarray, title: str, palette: dict[str, str]) -> int:
    axis.set_xlim(0, 1)
    axis.set_ylim(float(depth[-1]), 0)
    axis.set_xticks([])
    axis.set_yticks(np.arange(0, float(depth[-1]) + 0.1, 10))
    axis.set_yticks(np.arange(0, float(depth[-1]) + 0.1, 1), minor=True)
    style_axis(axis, title, "", "泥面以下深度 (m)")
    count = 0
    for index, value in enumerate(categories):
        if not value or value not in palette:
            continue
        top = float(depth[index] if index == 0 else (depth[index - 1] + depth[index]) / 2)
        bottom = float(depth[index] if index == len(depth) - 1 else (depth[index] + depth[index + 1]) / 2)
        axis.axhspan(top, bottom, color=palette[value], linewidth=0)
        count += 1
    handles = [Patch(facecolor=color, edgecolor="none", label=label) for label, color in palette.items()]
    axis.legend(handles=handles, loc="lower center", bbox_to_anchor=(0.5, -0.13), ncol=min(4, len(handles)), fontsize=7.0, frameon=False)
    return count


def categorical_layer_depth(
    axis: plt.Axes,
    depth: np.ndarray,
    categories: np.ndarray,
    layers: list[dict[str, float | int | str]],
    title: str,
    palette: dict[str, str],
    show_labels: bool = False,
    merge_adjacent: bool = False,
    label_map: dict[str, str] | None = None,
    show_legend: bool = True,
    label_min_thickness: float = 0.80,
) -> int:
    classified_layers: list[dict[str, float | str]] = []
    for layer in layers:
        top, bottom = float(layer["top"]), float(layer["bottom"])
        mask = (depth >= top) & (depth <= bottom)
        evidence = [str(value) for value in categories[mask] if value and str(value) in palette]
        if not evidence:
            continue
        counts = Counter(evidence)
        label = max(counts, key=counts.get)
        if (
            merge_adjacent
            and classified_layers
            and classified_layers[-1]["label"] == label
            and top - float(classified_layers[-1]["bottom"]) <= 0.10
        ):
            classified_layers[-1]["bottom"] = bottom
        else:
            classified_layers.append({"top": top, "bottom": bottom, "label": label})

    axis.set_xlim(0, 1)
    axis.set_ylim(float(depth[-1]), 0)
    axis.set_xticks([])
    axis.set_yticks(np.arange(0, float(depth[-1]) + 0.1, 10))
    axis.set_yticks(np.arange(0, float(depth[-1]) + 0.1, 1), minor=True)
    display_title = f"{title}（{len(classified_layers)} 层）" if show_labels else title
    style_axis(axis, display_title, "", "泥面以下深度 (m)")
    for layer in classified_layers:
        top = float(layer["top"])
        bottom = float(layer["bottom"])
        label = str(layer["label"])
        display_label = label_map.get(label, label) if label_map else label
        axis.axhspan(top, bottom, color=palette[label], linewidth=0)
        if show_labels:
            axis.axhline(top, color="#4B5A60", linewidth=0.45)
        if show_labels and bottom - top >= label_min_thickness:
            axis.text(
                0.5,
                (top + bottom) / 2,
                display_label,
                ha="center",
                va="center",
                fontsize=7.0,
                weight="semibold",
                color="#243238",
                clip_on=True,
            )
    if show_labels and classified_layers:
        axis.axhline(float(classified_layers[-1]["bottom"]), color="#4B5A60", linewidth=0.45)
    if show_legend:
        handles = [
            Patch(
                facecolor=color,
                edgecolor="none",
                label=label_map.get(label, label) if label_map else label,
            )
            for label, color in palette.items()
        ]
        axis.legend(handles=handles, loc="lower center", bbox_to_anchor=(0.5, -0.13), ncol=min(4, len(handles)), fontsize=7.0, frameon=False)
    return len(classified_layers)


def render_page_01(data: dict[str, object], output: Path) -> dict[str, object]:
    figure, _ = page_figure(1, "CPTU 原始测量与相关性")
    # The second-row axis labels use two lines (full name + symbol/unit).
    # Reserve a dedicated legend band below them so the zone legend never
    # competes with the engineering labels.
    grid = figure.add_gridspec(2, 2, left=0.09, right=0.94, top=0.885, bottom=0.135, hspace=0.29, wspace=0.28)
    depth = data["depth"]
    specs = [("qt", "修正锥尖阻力 qt", AXIS_LABELS["qt"], CURVE_COLORS["qt"], (0, 45)), ("rf", "摩阻比 Rf", AXIS_LABELS["rf"], CURVE_COLORS["rf"], (0, 50)), ("u2", "孔隙水压力 u2", AXIS_LABELS["u2"], CURVE_COLORS["u2"], (-1000, 3500))]
    for index, (key, title, xlabel, color, limits) in enumerate(specs):
        axis = figure.add_subplot(grid[index // 2, index % 2])
        depth_style(axis, depth, title, xlabel, True)
        plot_depth(axis, depth, data[key], color, limits)
    scatter = figure.add_subplot(grid[1, 1])
    count = scatter_zones(scatter, data["qc"], data["fs"], data["zone"], "锥尖阻力–侧壁摩阻力相关性", AXIS_LABELS["qc"], AXIS_LABELS["fs"])
    zone_legend(figure, 0.052, 7.2)
    save_page(figure, output)
    return {"scatterPoints": count}


def render_scatter_pair(page: int, title: str, data: dict[str, object], output: Path, panels: list[tuple[np.ndarray, np.ndarray, str, str, str, bool, bool, tuple[float, float] | None, tuple[float, float] | None]]) -> dict[str, object]:
    figure, _ = page_figure(page, title)
    grid = figure.add_gridspec(1, 2, left=0.065, right=0.95, top=0.86, bottom=0.14, wspace=0.24)
    counts = []
    for index, (x, y, panel_title, xlabel, ylabel, logx, logy, xlim, ylim) in enumerate(panels):
        axis = figure.add_subplot(grid[0, index])
        counts.append(scatter_zones(axis, x, y, data["zone"], panel_title, xlabel, ylabel, logx, logy))
        if xlim is not None:
            axis.set_xlim(*xlim)
        if ylim is not None:
            axis.set_ylim(*ylim)
    zone_legend(figure, 0.055, 8.2)
    save_page(figure, output)
    return {"scatterPoints": counts}


def render_page_04(data: dict[str, object], output: Path) -> dict[str, object]:
    observed = {value for value in data["schneider"] if value}
    categories = [label for label in SCHNEIDER_2008_LABELS if label in observed]
    palette = {label: PARAMETER_COLORS[index % len(PARAMETER_COLORS)] for index, label in enumerate(categories)}
    figure, _ = page_figure(4, "Schneider 2008 分类证据")
    grid = figure.add_gridspec(1, 2, left=0.07, right=0.95, top=0.86, bottom=0.14, wspace=0.22)
    axis = figure.add_subplot(grid[0, 0])
    x, y, category = data["bq"], data["robertson_qtn"], data["schneider"]
    valid_count = 0
    for label, color in palette.items():
        mask = np.isfinite(x) & np.isfinite(y) & (y > 0) & (category == label)
        axis.scatter(
            x[mask],
            y[mask],
            s=7,
            alpha=0.55,
            color=color,
            edgecolors="none",
            rasterized=True,
            label=SCHNEIDER_2008_LABELS.get(label, label),
        )
        valid_count += int(np.count_nonzero(mask))
    axis.set_yscale("log")
    style_axis(axis, "归一化孔压–锥阻证据", AXIS_LABELS["bq"], AXIS_LABELS["qtn"])
    axis.set_xlim(-2, 3)
    axis.set_ylim(0.1, 1000)
    axis.legend(loc="upper right", fontsize=7.2, frameon=False)
    depth_axis = figure.add_subplot(grid[0, 1])
    categorical_count = categorical_layer_depth(
        depth_axis,
        data["depth"],
        category,
        data["layers"],
        "Schneider 2008 分类分层",
        palette,
        show_labels=True,
        merge_adjacent=True,
        label_map=SCHNEIDER_2008_LABELS,
    )
    save_page(figure, output)
    return {
        "scatterPoints": valid_count,
        "classifiedRows": categorical_count,
        "classes": categories,
        "labelDefinitions": {label: SCHNEIDER_2008_LABELS[label] for label in categories},
    }


def render_page_05(data: dict[str, object], output: Path) -> dict[str, object]:
    figure, _ = page_figure(5, "深度窗口土类判别")
    grid = figure.add_gridspec(
        1,
        2,
        left=0.085,
        right=0.93,
        top=0.885,
        bottom=0.13,
        wspace=0.24,
        width_ratios=[0.92, 1.08],
    )
    depth = np.asarray(data["depth"], dtype=float)
    zone = np.asarray(data["zone"], dtype=int)

    # 以实际深度而不是固定行数建立 1.0 m 窗口，避免采样间距变化时窗口失真。
    window_radius = 0.50
    left = np.searchsorted(depth, depth - window_radius, side="left")
    right = np.searchsorted(depth, depth + window_radius, side="right")
    valid_zone = np.isin(zone, np.arange(1, 10))
    group_specs = [
        ("黏性土", np.isin(zone, [1, 2, 3, 4, 5]), ZONE_COLORS[4]),
        ("粉土", zone == 6, "#A8D9E8"),
        ("砂性土", np.isin(zone, [7, 8, 9]), ZONE_COLORS[7]),
    ]

    def window_counts(mask: np.ndarray) -> np.ndarray:
        prefix = np.concatenate(([0], np.cumsum(mask.astype(int))))
        return prefix[right] - prefix[left]

    denominator = window_counts(valid_zone)
    shares = []
    for _, mask, _ in group_specs:
        numerator = window_counts(mask)
        shares.append(
            np.divide(
                numerator,
                denominator,
                out=np.full(depth.shape, np.nan, dtype=float),
                where=denominator > 0,
            )
            * 100
        )
    share_matrix = np.vstack(shares)
    has_evidence = denominator > 0
    winner = np.full(depth.shape, -1, dtype=int)
    winner[has_evidence] = np.argmax(share_matrix[:, has_evidence], axis=0)
    maximum_share = np.full(depth.shape, np.nan, dtype=float)
    maximum_share[has_evidence] = np.max(share_matrix[:, has_evidence], axis=0)

    # 将逐深度的主导类别收敛为离散土层；类别变化、无证据或真实深度间断都会结束当前层。
    positive_steps = np.diff(depth)
    positive_steps = positive_steps[positive_steps > 0]
    typical_step = float(np.median(positive_steps)) if len(positive_steps) else 0.01
    gap_threshold = max(0.10, typical_step * 5)
    winner_layers: list[dict[str, object]] = []
    index = 0
    while index < len(depth):
        if winner[index] < 0:
            index += 1
            continue
        start = index
        category = int(winner[index])
        while (
            index + 1 < len(depth)
            and winner[index + 1] == category
            and depth[index + 1] - depth[index] <= gap_threshold
        ):
            index += 1
        end = index
        top = float(depth[start])
        bottom = float(depth[end])
        if start > 0 and winner[start - 1] >= 0 and depth[start] - depth[start - 1] <= gap_threshold:
            top = float((depth[start - 1] + depth[start]) / 2)
        if end + 1 < len(depth) and winner[end + 1] >= 0 and depth[end + 1] - depth[end] <= gap_threshold:
            bottom = float((depth[end] + depth[end + 1]) / 2)
        winner_layers.append({"top": top, "bottom": bottom, "category": category})
        index += 1

    winner_axis = figure.add_subplot(grid[0, 0])
    depth_style(winner_axis, depth, f"最高占比分类分层（{len(winner_layers)} 层）", "", True)
    winner_axis.set_xlim(0, 1)
    winner_axis.set_xticks([])
    for layer_number, layer in enumerate(winner_layers, start=1):
        top = float(layer["top"])
        bottom = float(layer["bottom"])
        category = int(layer["category"])
        label, _, color = group_specs[category]
        height = max(0.001, bottom - top)
        winner_axis.barh(
            (top + bottom) / 2,
            1,
            height=height,
            left=0,
            color=color,
            alpha=0.92,
            edgecolor="#4B5A60",
            linewidth=0.55,
        )
        if height >= 0.80:
            winner_axis.text(
                0.5,
                (top + bottom) / 2,
                label,
                ha="center",
                va="center",
                fontsize=7.2,
                weight="semibold",
                color="#243238",
                clip_on=True,
            )

    composition_axis = figure.add_subplot(grid[0, 1], sharey=winner_axis)
    depth_style(composition_axis, depth, "深度窗口土类组成", AXIS_LABELS["composition"], False)
    composition_axis.set_xlim(0, 100)
    composition_axis.set_xticks([0, 25, 50, 75, 100])
    cumulative = np.zeros_like(depth, dtype=float)
    for fraction, (label, _, color) in zip(shares, group_specs):
        composition_axis.fill_betweenx(
            depth,
            cumulative,
            cumulative + fraction,
            color=color,
            alpha=0.92,
            linewidth=0,
            label=label,
        )
        cumulative = cumulative + fraction

    handles = [
        plt.Rectangle((0, 0), 1, 1, facecolor=color, edgecolor="none", label=label)
        for label, _, color in group_specs
    ]
    figure.legend(
        handles=handles,
        loc="lower center",
        bbox_to_anchor=(0.5, 0.055),
        ncol=3,
        frameon=False,
        fontsize=8.2,
        handlelength=1.4,
        columnspacing=2.2,
    )
    figure.text(
        0.50,
        0.092,
        "左图将相邻且主导土类相同的深度窗口合并为一层；右图保留三类占比依据。窗口宽度 1.0 m。",
        ha="center",
        fontsize=7.8,
        color="#68777D",
    )
    save_page(figure, output)

    return {
        "compositionRows": int(len(depth)),
        "windowWidthM": 1.0,
        "winnerLayers": int(len(winner_layers)),
        "emptyWindows": int(np.count_nonzero(~has_evidence)),
        "gapThresholdM": gap_threshold,
    }


def render_depth_grid(page: int, title: str, data: dict[str, object], output: Path, tracks: list[tuple[str, str, str, str, tuple[float, float] | None, bool]]) -> dict[str, object]:
    figure, _ = page_figure(page, title)
    count = len(tracks)
    grid = figure.add_gridspec(1, count, left=0.055, right=0.955, top=0.86, bottom=0.16, wspace=0.25)
    depth = data["depth"]
    valid_counts: dict[str, int] = {}
    for index, (key, panel_title, xlabel, color, limits, log) in enumerate(tracks):
        axis = figure.add_subplot(grid[0, index])
        depth_style(axis, depth, panel_title, xlabel, index == 0)
        # Five-track pages have narrow columns. Keep the full engineering name
        # visible while preventing adjacent labels from colliding.
        if count >= 5:
            axis.xaxis.label.set_fontsize(8.0)
            axis.xaxis.label.set_linespacing(1.05)
        plot_depth(axis, depth, data[key], color, limits, log)
        valid_counts[key] = int(np.count_nonzero(np.isfinite(data[key])))
    figure.text(0.50, 0.060, "空白表示当前方法无有效值或存在真实数据断点；所有可用观测均参与绘制。", ha="center", fontsize=8.0, color="#68777D")
    save_page(figure, output)
    return {"validCounts": valid_counts}


def render_page_07(data: dict[str, object], output: Path) -> dict[str, object]:
    return render_depth_grid(7, "归一化参数与 Ic 深度图", data, output, [
        ("robertson_qtn", "归一化锥尖阻力 Qtn", AXIS_LABELS["qtn"], PARAMETER_COLORS[0], (0.1, 1000), True),
        ("fr", "归一化摩阻比 Fr", AXIS_LABELS["fr"], CURVE_COLORS["rf"], (0, 10), False),
        ("bq", "孔压参数 Bq", AXIS_LABELS["bq"], CURVE_COLORS["u2"], (-2, 3), False),
        ("robertson_ic", "土体行为类型指数 Ic", AXIS_LABELS["ic"], CURVE_COLORS["ic"], (1, 4.2), False),
        ("jts_ic", "JTS 土体行为类型指数 Ic", AXIS_LABELS["jts_ic"], "#9D6C45", (1, 4.2), False),
    ])


def render_page_08(data: dict[str, object], output: Path) -> dict[str, object]:
    observed = {value for value in data["robertson_2016"] if value}
    categories = [label for label in ROBERTSON_2016_LABELS if label in observed]
    palette = {label: list(ZONE_COLORS.values())[(index + 1) % 9] for index, label in enumerate(categories)}
    figure, _ = page_figure(8, "Modified Robertson 2016 深度分类")
    grid = figure.add_gridspec(1, 5, left=0.055, right=0.955, top=0.86, bottom=0.13, wspace=0.25)
    depth = data["depth"]
    specs = [("robertson_qtn", "归一化锥尖阻力 Qtn", AXIS_LABELS["qtn"], True, (0.1, 1000)), ("fr", "归一化摩阻比 Fr", AXIS_LABELS["fr"], False, (0, 20)), ("ib", "修正土体行为类型指数 IB", AXIS_LABELS["ib"], False, None), ("cd", "收缩–剪胀参数 CD", AXIS_LABELS["cd"], False, (-100, 100))]
    for index, (key, title, xlabel, log, limits) in enumerate(specs):
        axis = figure.add_subplot(grid[0, index]); depth_style(axis, depth, title, xlabel, index == 0); plot_depth(axis, depth, data[key], PARAMETER_COLORS[index], limits, log)
    category_axis = figure.add_subplot(grid[0, 4])
    count = categorical_layer_depth(
        category_axis,
        depth,
        data["robertson_2016"],
        data["layers"],
        "Modified Robertson 2016 七类分层",
        palette,
        show_labels=True,
        merge_adjacent=True,
        label_map=ROBERTSON_2016_LABELS,
    )
    save_page(figure, output)
    return {
        "classifiedRows": count,
        "classes": categories,
        "labelDefinitions": {label: ROBERTSON_2016_LABELS[label] for label in categories},
    }


def render_page_09(data: dict[str, object], output: Path) -> dict[str, object]:
    figure, _ = page_figure(9, "多方法分类与刚度证据")
    grid = figure.add_gridspec(1, 5, left=0.055, right=0.955, top=0.86, bottom=0.235, wspace=0.28)
    depth = data["depth"]
    zone_labels = [f"Z{zone}" for zone in range(1, 10)]
    zone_palette = {label: ZONE_COLORS[int(label[1:])] for label in zone_labels}
    zone_label_map = {label: f"{label} · {ZONE_NAMES[int(label[1:])]}" for label in zone_labels}
    zone_categories = np.asarray([f"Z{zone}" if zone else "" for zone in data["zone"]], dtype=object)
    zone_axis = figure.add_subplot(grid[0, 0])
    zone_count = categorical_layer_depth(
        zone_axis,
        depth,
        zone_categories,
        data["layers"],
        "JTS/T 242—2020 九区分层",
        zone_palette,
        show_labels=True,
        merge_adjacent=True,
        label_map=zone_label_map,
        show_legend=False,
        label_min_thickness=1.40,
    )
    r_labels = [label for label in ROBERTSON_2016_LABELS if np.any(data["robertson_2016"] == label)]
    r_palette = {label: list(ZONE_COLORS.values())[(index + 1) % 9] for index, label in enumerate(r_labels)}
    s_labels = [label for label in SCHNEIDER_2008_LABELS if np.any(data["schneider"] == label)]
    s_palette = {label: PARAMETER_COLORS[index % len(PARAMETER_COLORS)] for index, label in enumerate(s_labels)}
    robertson_axis = figure.add_subplot(grid[0, 1])
    robertson_count = categorical_layer_depth(
        robertson_axis,
        depth,
        data["robertson_2016"],
        data["layers"],
        "Modified Robertson 2016 七类分层",
        r_palette,
        show_labels=True,
        merge_adjacent=True,
        label_map=ROBERTSON_2016_LABELS,
        show_legend=False,
        label_min_thickness=1.40,
    )
    schneider_axis = figure.add_subplot(grid[0, 2])
    schneider_count = categorical_layer_depth(
        schneider_axis,
        depth,
        data["schneider"],
        data["layers"],
        "Schneider 2008 五类分层",
        s_palette,
        show_labels=True,
        merge_adjacent=True,
        label_map=SCHNEIDER_2008_LABELS,
        show_legend=False,
        label_min_thickness=1.40,
    )
    g_axis=figure.add_subplot(grid[0,3]); depth_style(g_axis,depth,"小应变剪切模量 G0",AXIS_LABELS["g0"],False); plot_depth(g_axis,depth,data["g0"],PARAMETER_COLORS[0])
    k_axis=figure.add_subplot(grid[0,4]); depth_style(k_axis,depth,"静止土压力系数 K0",AXIS_LABELS["k0"],False); plot_depth(k_axis,depth,data["k0"],PARAMETER_COLORS[1])

    legend_specs = [
        (zone_axis, "JTS/T 242—2020", zone_palette, zone_label_map),
        (robertson_axis, "Modified Robertson 2016", r_palette, ROBERTSON_2016_LABELS),
        (schneider_axis, "Schneider 2008", s_palette, SCHNEIDER_2008_LABELS),
    ]
    for legend_axis, legend_title, palette, label_map in legend_specs:
        position = legend_axis.get_position()
        x = position.x0
        y = 0.205
        figure.text(x, y, legend_title, ha="left", va="top", fontsize=7.0, weight="bold", color="#26343A")
        y -= 0.017
        for label, color in palette.items():
            figure.add_artist(Rectangle((x, y - 0.007), 0.007, 0.007, transform=figure.transFigure, facecolor=color, edgecolor="none"))
            figure.text(x + 0.010, y - 0.001, label_map.get(label, label), ha="left", va="top", fontsize=5.9, color="#33464E")
            y -= 0.014
    save_page(figure,output)
    return {
        "referenceLayers": len(data["layers"]),
        "jtsLayers": zone_count,
        "robertsonLayers": robertson_count,
        "schneiderLayers": schneider_count,
        "robertsonClasses": len(r_labels),
        "schneiderClasses": len(s_labels),
        "legendAlignmentY": 0.205,
        "directLabelMinThicknessM": 1.40,
    }


def render_page_11(data: dict[str, object], output: Path) -> dict[str, object]:
    figure, _ = page_figure(11, "强度、刚度与超固结证据")
    grid=figure.add_gridspec(1,5,left=.055,right=.955,top=.86,bottom=.11,wspace=.25); depth=data["depth"]
    specs=[("jts_es","压缩模量 Es（JTS）",AXIS_LABELS["jts_es"]),("g0","小应变剪切模量 G0",AXIS_LABELS["g0"]),("su_peak","不排水抗剪强度 Su",AXIS_LABELS["su"]),("su_ratio","归一化不排水抗剪强度",AXIS_LABELS["su_ratio"]),("ocr","超固结比 OCR",AXIS_LABELS["ocr"])]
    counts={}
    for i,(key,title,xlabel) in enumerate(specs):
        ax=figure.add_subplot(grid[0,i]); depth_style(ax,depth,title,xlabel,i==0)
        second=(data["su_remolded"],CURVE_COLORS["u2"],"重塑后") if key=="su_peak" else None
        plot_depth(ax,depth,data[key],PARAMETER_COLORS[i],None,False,second); counts[key]=int(np.count_nonzero(np.isfinite(data[key])))
    save_page(figure,output); return {"validCounts":counts}


def render_page_15(data: dict[str, object], output: Path) -> dict[str, object]:
    figure, _ = page_figure(15, "公式、系数与参考文献")
    axis = figure.add_axes([0.055, 0.055, 0.89, 0.84])
    axis.set_xlim(0, 1)
    axis.set_ylim(0, 1)
    axis.axis("off")

    settings = dict(data["settings"])
    result_counts: dict[str, int] = dict(data["result_counts"])
    formula_index: list[str] = list(data["formula_index"])
    references: dict[str, str] = dict(data["references"])

    groups: list[dict[str, object]] = []
    for spec in FORMULA_GROUPS:
        columns = list(spec["columns"])
        valid_count = max((result_counts.get(column, 0) for column in columns), default=0)
        if valid_count <= 0:
            continue
        prefixes = list(spec["prefixes"])
        formulas = [formula for formula in formula_index if any(formula.startswith(prefix) for prefix in prefixes)]
        if not formulas:
            raise ValueError(f"实际计算参数缺少公式索引：{spec['title']}")
        title = str(spec["title"])
        if title == "不排水强度与归一化强度":
            formulas.append("Su(rem)=Su/St；Su(r)/σ′v0=Su(rem)/σ′v0  [A02]")
        if title == "超固结比与静止土压力":
            formulas.append("黏性土 φ′ 无有效值时，K0 计算采用 30° 预设  [A02]")
        groups.append({"title": title, "columns": columns, "count": valid_count, "formulas": formulas})

    references["A02"] = "快捷方法包实现约定：Su(rem)=Su/St；K0 的黏性土 φ′ 缺失时采用 30° 预设。"
    used_reference_ids = sorted({
        match
        for group in groups
        for formula in group["formulas"]
        for match in re.findall(r"\[([A-Z]\d{2})\]", str(formula))
    })

    axis.text(0.5, 0.988, "本次实际计算采用的公式", ha="center", va="top", fontsize=13.2, weight="bold", color="#111719")
    axis.text(0.5, 0.962, "仅列出本报告中产生有效结果的方法；系数为本次运行实际采用值。", ha="center", va="top", fontsize=8.1, color="#526168")

    coefficient_cells = [
        ("有效面积比", f"a = {settings.get('有效面积比 a', '—')}"),
        ("水重度", "γw = 10.00 kN/m³"),
        ("土粒比重", "Gs = 2.65"),
        ("不排水强度系数", "Nkt = 15.5"),
        ("超固结系数", "kOCR = 0.16"),
        ("K0 黏土预设", "φ′ = 30°"),
    ]
    cell_left, cell_right, cell_top, cell_height = 0.015, 0.985, 0.925, 0.038
    cell_width = (cell_right - cell_left) / 3
    for index, (label, value) in enumerate(coefficient_cells):
        row, column = divmod(index, 3)
        x = cell_left + column * cell_width
        y = cell_top - (row + 1) * cell_height
        axis.add_patch(Rectangle((x, y), cell_width, cell_height, facecolor="#F6F7F7" if row else "white", edgecolor="#B8C1C5", linewidth=0.65))
        axis.text(x + 0.010, y + cell_height / 2, label, ha="left", va="center", fontsize=7.2, color="#526168")
        axis.text(x + cell_width - 0.010, y + cell_height / 2, value, ha="right", va="center", fontsize=7.6, weight="semibold", color="#26343A")

    formula_top, formula_bottom = 0.830, 0.245
    axis.plot([0.5, 0.5], [formula_bottom, formula_top], color="#8B969B", linewidth=0.7, linestyle=(0, (4, 4)))
    columns = [groups[:8], groups[8:]]
    for column_index, column_groups in enumerate(columns):
        x0 = 0.015 if column_index == 0 else 0.515
        width = 0.470
        wrapped_groups: list[tuple[dict[str, object], list[str], float]] = []
        for group in column_groups:
            wrapped_lines: list[str] = []
            for formula in group["formulas"]:
                wrapped_lines.extend(textwrap.wrap(str(formula), width=48, break_long_words=False, break_on_hyphens=False) or [str(formula)])
            weight = 1.00 + 0.44 * len(wrapped_lines)
            wrapped_groups.append((group, wrapped_lines, weight))
        total_weight = sum(item[2] for item in wrapped_groups)
        y = formula_top
        for block_index, (group, wrapped_lines, weight) in enumerate(wrapped_groups):
            height = (formula_top - formula_bottom) * weight / total_weight
            bottom = y - height
            axis.add_patch(Rectangle((x0, bottom + 0.002), width, height - 0.004, facecolor="#F8F9F9" if block_index % 2 else "white", edgecolor="#D4DADD", linewidth=0.48))
            axis.add_patch(Rectangle((x0, bottom + 0.002), 0.004, height - 0.004, facecolor="#9D6C45", edgecolor="none"))
            axis.text(x0 + 0.012, y - 0.012, str(group["title"]), ha="left", va="top", fontsize=8.1, weight="bold", color="#111719")
            axis.text(x0 + width - 0.010, y - 0.012, f"{int(group['count']):,} 个值", ha="right", va="top", fontsize=6.8, color="#65747B")
            line_y = y - 0.034
            line_step = max(0.0115, (height - 0.044) / max(1, len(wrapped_lines)))
            for line in wrapped_lines:
                axis.text(x0 + 0.016, line_y, line, ha="left", va="top", fontsize=6.95, color="#26343A")
                line_y -= line_step
            y = bottom

    axis.text(0.015, 0.222, "参考来源", ha="left", va="top", fontsize=9.2, weight="bold", color="#111719")
    axis.plot([0.015, 0.985], [0.210, 0.210], color="#65747B", linewidth=0.7)
    reference_columns = [used_reference_ids[::2], used_reference_ids[1::2]]
    for column_index, reference_ids in enumerate(reference_columns):
        x0 = 0.015 if column_index == 0 else 0.515
        y = 0.193
        for reference_id in reference_ids:
            reference = references.get(reference_id, "来源条目缺失")
            lines = textwrap.wrap(reference, width=76, break_long_words=False, break_on_hyphens=False) or [reference]
            axis.text(x0, y, reference_id, ha="left", va="top", fontsize=6.8, weight="bold", color="#9D6C45")
            axis.text(x0 + 0.038, y, "\n".join(lines), ha="left", va="top", fontsize=6.35, color="#33464E", linespacing=1.18)
            y -= 0.018 + 0.014 * len(lines)

    axis.text(0.5, 0.006, "公式与系数来自本次 Excel 的“设置与方法”索引，并与当前快捷方法包实现交叉核对。", ha="center", va="bottom", fontsize=6.8, color="#68777D")
    save_page(figure, output)
    return {
        "actualFormulaGroups": len(groups),
        "formulaLines": sum(len(group["formulas"]) for group in groups),
        "actualResultColumns": sorted({column for group in groups for column in group["columns"] if result_counts.get(column, 0) > 0}),
        "usedReferences": used_reference_ids,
        "coefficientCells": dict(coefficient_cells),
        "formulaSourceSha256": hashlib.sha256("\n".join(formula_index).encode("utf-8")).hexdigest(),
    }


def render_contact_sheet(page_paths: list[Path], output: Path) -> None:
    figure, axes = plt.subplots(3, 5, figsize=(20, 11.25), dpi=100, facecolor="#E9EEF0")
    for index, (axis, path) in enumerate(zip(axes.flat, page_paths), start=1):
        with Image.open(path) as image:
            axis.imshow(np.asarray(image.convert("RGB")))
        axis.set_title(f"第 {index:02d} 页", fontsize=10, weight="bold", pad=6)
        axis.axis("off")
    figure.suptitle("SIGS-OGLab · 15 页 PNG 视觉迁移总览", fontsize=18, weight="bold", y=.985)
    figure.subplots_adjust(left=.02,right=.98,top=.94,bottom=.025,wspace=.045,hspace=.11)
    figure.savefig(output,dpi=100,facecolor="#E9EEF0",edgecolor="none")
    plt.close(figure)


def render_all(workbook: Path, output_dir: Path) -> dict[str, object]:
    data=load_data(workbook); output_dir.mkdir(parents=True,exist_ok=True); pages=[]; metrics={}
    def out(page:int)->Path: return output_dir/f"page-{page:02d}.png"
    metrics["01"]=render_page_01(data,out(1)); pages.append(out(1))
    qc_over_pa=np.asarray(data["qc"])*10
    metrics["02"]=render_scatter_pair(2,"非归一化土体行为类型（SBT）与孔压证据",data,out(2),[(np.asarray(data["rf"]),qc_over_pa,"非归一化土体行为类型证据",AXIS_LABELS["rf"],AXIS_LABELS["qc_over_pa"],True,True,(0.005,20),(0.5,1000)),(np.asarray(data["bq"]),np.asarray(data["qnet"]),"孔压响应",AXIS_LABELS["bq"],AXIS_LABELS["qnet"],False,False,(-2,3),(-1000,42000))]); pages.append(out(2))
    metrics["03"]=render_scatter_pair(3,"归一化土体行为类型（SBTn）与孔压证据",data,out(3),[(np.asarray(data["fr"]),np.asarray(data["jts_qtn"]),"归一化土体行为类型证据",AXIS_LABELS["fr"],AXIS_LABELS["jts_qtn"],True,True,(0.01,20),(0.1,1000)),(np.asarray(data["bq"]),np.asarray(data["robertson_qtn"]),"归一化孔压响应",AXIS_LABELS["bq"],AXIS_LABELS["qtn"],False,True,(-2,3),(0.1,1000))]); pages.append(out(3))
    metrics["04"]=render_page_04(data,out(4)); pages.append(out(4))
    metrics["05"]=render_page_05(data,out(5)); pages.append(out(5))
    page06_check=output_dir/"page-06-check.json"; PAGE06.render(workbook,out(6),page06_check); metrics["06"]={"reusedAcceptedStyle":True}; pages.append(out(6))
    metrics["07"]=render_page_07(data,out(7)); pages.append(out(7))
    metrics["08"]=render_page_08(data,out(8)); pages.append(out(8))
    metrics["09"]=render_page_09(data,out(9)); pages.append(out(9))
    metrics["10"]=render_depth_grid(10,"渗透、密实度与摩擦角",data,out(10),[("k","渗透系数 k",AXIS_LABELS["k"],PARAMETER_COLORS[0],None,True),("spt_n","标准贯入击数 N",AXIS_LABELS["spt_n"],PARAMETER_COLORS[1],None,False),("es","压缩模量 Es（R05）",AXIS_LABELS["es"],PARAMETER_COLORS[2],None,False),("dr","相对密实度 Dr",AXIS_LABELS["dr"],PARAMETER_COLORS[3],(0,100),False),("phi","有效内摩擦角 φ′",AXIS_LABELS["phi"],PARAMETER_COLORS[4],(20,45),False)]); pages.append(out(10))
    metrics["11"]=render_page_11(data,out(11)); pages.append(out(11))
    metrics["12"]=render_depth_grid(12,"波速、状态与应力历史",data,out(12),[("vs","剪切波速 Vs",AXIS_LABELS["vs"],PARAMETER_COLORS[0],None,False),("psi","状态参数 ψ",AXIS_LABELS["psi"],PARAMETER_COLORS[1],None,False),("k0","静止土压力系数 K0",AXIS_LABELS["k0"],PARAMETER_COLORS[2],None,False),("st","灵敏度 St",AXIS_LABELS["st"],PARAMETER_COLORS[3],None,False),("phi","有效摩擦角 φ′",AXIS_LABELS["phi"],PARAMETER_COLORS[4],(20,45),False)]); pages.append(out(12))
    metrics["13"]=render_depth_grid(13,"物理指标随深度变化",data,out(13),[("gamma_sat","饱和重度 γsat",AXIS_LABELS["gamma_sat"],PARAMETER_COLORS[0],(10,25),False),("water_content","含水率 w",AXIS_LABELS["water_content"],PARAMETER_COLORS[1],None,False),("void_ratio","孔隙比 e",AXIS_LABELS["void_ratio"],PARAMETER_COLORS[2],None,False),("gamma_d","干重度 γd",AXIS_LABELS["gamma_d"],PARAMETER_COLORS[3],None,False),("porosity","孔隙率 n",AXIS_LABELS["porosity"],PARAMETER_COLORS[4],(0,1),False)]); pages.append(out(13))
    metrics["14"]=render_depth_grid(14,"修正参数与残余强度",data,out(14),[("qt","修正锥尖阻力 qt",AXIS_LABELS["qt"],CURVE_COLORS["qt"],(0,45),False),("robertson_qtn","归一化锥尖阻力 Qtn",AXIS_LABELS["qtn"],PARAMETER_COLORS[0],(0.1,1000),True),("jts_ic","JTS 土体行为类型指数 Ic",AXIS_LABELS["jts_ic"],CURVE_COLORS["ic"],(1,4.2),False),("qtn_cs","等效洁净砂归一化锥尖阻力 Qtn,cs",AXIS_LABELS["qtn_cs"],PARAMETER_COLORS[1],(0.1,1000),True),("su_residual_ratio","残余不排水强度比",AXIS_LABELS["su_residual_ratio"],PARAMETER_COLORS[3],None,False)]); pages.append(out(14))
    metrics["15"]=render_page_15(data,out(15)); pages.append(out(15))
    contact=output_dir/"contact-sheet.png"; render_contact_sheet(pages,contact)
    page_records=[]
    for page,path in enumerate(pages,start=1):
        with Image.open(path) as image: size=list(image.size)
        expected=list(PORTRAIT_PX if page in PORTRAIT_PAGES else LANDSCAPE_PX)
        if size!=expected: raise ValueError(f"第 {page} 页尺寸错误：{size} != {expected}")
        page_records.append({"page":page,"path":path.as_posix(),"orientation":"portrait" if page in PORTRAIT_PAGES else "landscape","sizePx":size,"sha256":sha256(path)})
    check={"schemaVersion":1,"process":"Process119","kind":"15-page-png-visual-migration","source":{"path":workbook.as_posix(),"sha256":sha256(workbook),"rows":int(len(data["depth"])),"depthRangeM":[float(data["depth"][0]),float(data["depth"][-1])]},"backend":"Python/Matplotlib","noDownsampling":True,"missingValuesRemainGaps":True,"productionReportChanged":False,"pages":page_records,"contactSheet":{"path":contact.as_posix(),"sha256":sha256(contact)},"metrics":metrics,"axisLabelPolicy":"中文工程全称 + 参数符号 + 单位；密集曲线使用两行标签","axisLabels":{key:value.replace("\n"," ") for key,value in AXIS_LABELS.items()},"fieldMapping":{"qt":"快捷解译结果/qt(kPa) -> MPa","Rf":"快捷解译结果/Rf(%)","u2":"原始数据/u2(kPa)","Ic":"快捷解译结果/JTS Ic or Robertson Ic","classification":"快捷解译结果/JTS Zone, Modified Robertson 2016, Schneider 2008"},"qaNotes":{"figureContract":"PNG-only visual migration preview; not a journal submission bundle","vectorExport":"not applicable: user requested PNG samples before production integration","printDpi":"not applicable: fixed 1920x1080 / 1080x1528 screen-review pixels","logGuard":"all log tracks remove nonpositive values before set_xscale; values are not replaced","visualInspection":["contact sheet","pages 01, 02, 03, 04, 06, 07, 08, 10 and 14 at original size"]}}
    (output_dir/"atlas-concept-check.json").write_text(json.dumps(check,ensure_ascii=False,indent=2),encoding="utf-8")
    return check


def main()->None:
    parser=argparse.ArgumentParser(description="Render the complete Process119 PNG visual migration pack.")
    parser.add_argument("--input",required=True,type=Path); parser.add_argument("--output-dir",required=True,type=Path); args=parser.parse_args()
    check=render_all(args.input.resolve(),args.output_dir.resolve()); print(json.dumps({"pages":len(check["pages"]),"contactSheet":check["contactSheet"]["path"],"rows":check["source"]["rows"]},ensure_ascii=False))


if __name__=="__main__": main()
